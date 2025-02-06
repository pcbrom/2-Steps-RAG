import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Define control variables
variaveis_controle = {
    'model': ['text-davinci-003', 'text-curie-001', 'text-babbage-001'],
    'temperature': [0.1, 0.5, 0.9],
    'top_p': [0.1, 0.5, 0.9],
    'top_k': [1, 5, 10]
}

# Create a pandas DataFrame
df = pd.DataFrame(variaveis_controle)

# Generate all possible combinations of control variables
df_completo = df.loc[df.index.repeat(len(df)**3)].reset_index(drop=True)

# Create a new DataFrame with all combinations
df_final = pd.DataFrame(df_completo.values, columns=['temperature', 'top_p', 'top_k', 'model'])

# Add a column for the prompt (to be filled later)
df_final['prompt'] = ''

# Add a column for the results (to be filled later)
df_final['results'] = ''

# Save the DataFrame to an XLSX file with good formatting
wb = Workbook()
ws = wb.active

# Write the header
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
for col_num, value in enumerate(df_final.columns.values):
    ws.cell(row=1, column=col_num+1, value=value).font = header_font
    ws.cell(row=1, column=col_num+1).fill = header_fill

# Write the data
for row in df_final.values.tolist():
    ws.append(row)

wb.save("experimental_design_plan.xlsx")